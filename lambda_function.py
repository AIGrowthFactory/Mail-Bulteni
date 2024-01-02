import datetime
import json
import os
import unicodedata
import ssl
import requests
from bs4 import BeautifulSoup
import openpyxl
import smtplib
import re
from datetime import date
import boto3
from email.message import EmailMessage



# Extraction rules for pazarlamasyon.com
pazarlamasyon_rules = {
    'date': {'class_': 'sp-date'},
    'title': {'class_': 'single-post-title'},
    'content': {'class_': 'binduz-er-text'},
    'date_format': 'dd/mm/yyyy'
}

# Extraction rules for gh.com
gh_rules = {
    'date': {'class_': 'published updated', 'attrs': {'datetime': True}},
    'title': {'class_': 'post-title post-item-title'},
    'content': {'name': 'p'},
    'date_format': 'mm/dd/yyyy'
}

# Extraction rules for swipeline.com
swipeline_rules = {
    'date': {'class_': 'frontpage-post-date'},
    'title': {'class_': 'entry-title'},
    'content': {'class_': 'entry-content'},
    'date_format': 'none'
}

# Extraction rules for egirisim.com
egirisim_rules = {
    'title': {'class_': 'tdb-title-text'},
    'content': {'class_': 'tdb-block-inner td-fix-index'},
    'date_extraction': lambda url: url[21:31] if len(url) >= 31 else "invalid URL",
    'date_format': 'yyyy/mm/dd'
}

# Extraction rules for webrazzi.com
webrazzi_rules = {
    'title': {'class_': 'single-post-title'},
    'content': {'class_': 'single-post-content'},
    'date_extraction': lambda url: url[21:31] if len(url) >= 31 else "invalid URL",
    'date_format': 'yyyy/mm/dd'
}


def download_file_from_s3(bucket_name, object_key, local_path):
    """
    Download a file from an S3 bucket to the local file system.

    Parameters:
        bucket_name (str): The name of the S3 bucket.
        object_key (str): The key of the S3 object (file) to download.
        local_path (str): The local file path where the S3 object will be downloaded.

    Returns:
        bool: True if the file was downloaded successfully, False otherwise.
    """
    s3 = boto3.client('s3')

    try:
        s3.download_file(bucket_name, object_key, local_path)
        print(f"File downloaded successfully to {local_path}")
        return True
    except Exception as e:
        print(f"Error downloading file: {e}")
        return False

def summarizer(text):
    print("sumarizerdayım"+text)
    url = "https://api.meaningcloud.com/summarization-1.0"
    
    payload={
        'key': 'd2746048e56d9563a99d1857ced03a6b',
        'txt': text,
        'sentences': '1'
    }
    
    response = requests.post(url, data=payload)
    print(response.json())
    
    return response.json()["summary"]


def upload_file_to_s3(file_path, bucket_name, object_key):
    """
    Upload a file to an S3 bucket.

    Parameters:
        file_path (str): The local path to the file you want to upload.
        bucket_name (str): The name of the S3 bucket where the file will be uploaded.
        object_key (str): The object key (path) under which the file will be stored in the bucket.
    """
    try:
        # Create an S3 client
        s3 = boto3.client('s3')

        # Upload the file to S3
        with open(file_path, 'rb') as file:
            s3.upload_fileobj(file, bucket_name, object_key)

        print(f"File '{file_path}' uploaded to S3 bucket '{bucket_name}' with object key '{object_key}'")
        return True
    except Exception as e:
        print(f"Error uploading file to S3: {e}")
        return False


def log_to_file(message, log_file_path="log.txt"):
    """
    Appends a log message with a timestamp to a text file.

    Parameters:
        message (str): The log message to be written.
        log_file_path (str, optional): The path to the log file (default: "log.txt").
    """
    # Get the current timestamp
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Create the log entry with timestamp
    log_entry = f"[{timestamp}] {message}\n"

    try:
        # Open the log file in append mode and write the log entry
        with open(log_file_path, "a") as log_file:
            log_file.write(log_entry)
    except Exception as e:
        # Handle any exceptions that may occur during file write
        print(f"Error writing to log file: {e}")


def convert_to_dd_mm_yyyy(date_string, date_format):
    # Remove any non-digit characters and split the date_string
    date_string = ''.join(filter(str.isdigit, date_string))

    if date_format == 'dd/mm/yyyy':
        return date(int(date_string[4:8]), int(date_string[2:4]), int(date_string[0:2]))
    elif date_format == 'mm/dd/yyyy':
        return date(int(date_string[4:8]), int(date_string[0:2]), int(date_string[2:4]))
    elif date_format == 'yyyy/mm/dd':
        return date(int(date_string[0:4]), int(date_string[4:6]), int(date_string[6:8]))

    return None  # Return None for unrecognized or invalid date formats


def remove_chars_before_first_number(input_string):
    # Use regular expression to find the first occurrence of a digit
    match = re.search(r'\d', input_string)

    if match:
        # Get the index of the first digit
        index_of_first_digit = match.start()
        # Remove characters before the first digit
        result_string = input_string[index_of_first_digit:]
        return result_string
    else:
        # No digit found, return the original string
        return input_string


def extract_content_from_website(url, extraction_rules):
    """
    Extract content from a website based on provided extraction rules.

    Parameters:
        url (str): The URL of the webpage to scrape.
        extraction_rules (dict): A dictionary containing extraction rules for the website.

    Returns:
        tuple: A tuple containing the extracted content, URL, date, and title.
    """
    content_list = []
    final_string = ""

    # Send an HTTP GET request to the URL
    response = requests.get(url)

    # Parse the HTML content using BeautifulSoup
    soup = BeautifulSoup(response.content, "html.parser")

    # Extract title based on extraction rules
    
    title = soup.find(**extraction_rules['title']).text
    print(title,url)
    #try:
     #   title = title.text
    #except:
     #   return 0,0,0,0

    # Attempt to extract date using the provided date extraction method
    if 'date_extraction' in extraction_rules:
        date_extraction_method = extraction_rules['date_extraction']
        date = date_transformer(date_extraction_method(url), extraction_rules['date_format'])
    else:
        # Fallback date extraction method: Try to find a date element in the page
        date_element = soup.find(**extraction_rules['date'])
        if date_element:
            date = date_transformer(date_element.text, extraction_rules['date_format'])
        else:
            date = "Date not found"

    # Extract content based on extraction rules
    content_elements = soup.find_all(**extraction_rules['content'])

    for element in content_elements:
        paragraph_text = element.get_text(strip=True) + " "
        content_list.append(paragraph_text)

    for i in content_list:
        final_string += i + " "

    return unicodedata.normalize("NFKD", final_string) + " ***", url, date, title


def general_news_links(url, my_set, html_class, tag=None, attributes=None, exclusion_list="!@#$", root_relative=False):
    """
    Retrieve and parse a webpage to extract news links.

    Parameters:
        url (str): The URL of the webpage to scrape.
        my_set (set): A set to store unique hrefs.
        html_class (str): The HTML class name to search for.
        tag (str, optional): The HTML tag name to search for (default: None).
        attributes (dict, optional): A dictionary of attributes to search for within elements (default: None).
        exclusion_list (str, optional): A string of characters that, if found in an href, excludes it (default: "!@#$").
        root_relative (bool, optional): Whether to make hrefs root-relative by prepending the URL (default: False).

    Returns:
        set: A set containing the unique extracted hrefs.
    """
    # Create an empty set to store unique hrefs
    hrefs = set()

    # Create an empty list to store HTML elements
    elements = []

    # Send an HTTP GET request to the URL
    response = requests.get(url)

    # Parse the HTML content using BeautifulSoup
    soup = BeautifulSoup(response.content, "html.parser")

    # Find all elements with the specified tag or class
    if tag is not None:
        elements = soup.find_all(tag)
    elements += soup.find_all(class_=html_class)

    # Extract and collect the unique URLs
    for element in elements:
        # If attributes are specified, find the inner element with those attributes
        if attributes is not None:
            element = element.find(attributes)

        # Check if the 'href' attribute exists in the element's attributes
        if 'href' in element.attrs:
            if root_relative:
                href = url + element['href']
            else:
                href = element['href']

            # Check if the href is not already in the provided set and not in the exclusion list
            if href not in my_set and exclusion_list not in href:
                # Otherwise, add the href as is to the set
                hrefs.add(href)

    # Print the number of unique hrefs found
    print(f"Total unique hrefs found: {len(hrefs), url}")

    return hrefs


def check_words_in_string(word_list, input_string):
    normalized_input = unicodedata.normalize("NFKD", input_string.lower())
    return all(" " + unicodedata.normalize("NFKD", word.lower()) in normalized_input for word in word_list)


def find_next_available_row(sheet):
    for row_idx, row in enumerate(sheet.iter_rows(), start=1):
        if all(cell.value is None for cell in row):
            return row_idx
    return sheet.max_row + 1

def selected_operator_formatting(selected_operator):
    if selected_operator == "AND" :
        return "ve"
    else:
        return "ya da"

def email_sender(matching_urls, selected_operator):
    sender_email = "yapayzekanewsletter@gmail.com"
    password = "pcwvhszsorkhcsxp"
    #reciever_emails = ["bartucelasun@gmail.com", "mcelasun20@ku.edu.tr", "bartucelasun@hotmail.com"]
    reciever_emails = ["ali.tuncgenc@trakyayatirim.com.tr","baris.karakullukcu@trakyayatirim.com.tr","atakanzky@gmail.com","melisa.cetin@trakyayatirim.com.tr","bartucelasun@gmail.com", "kadir.bulut@trakyayatirim.com.tr", "kadirbul@gmail.com", "yasemin.tavlasoglu@trakyayatirim.com.tr", "erhan.petek@trakyayatirim.com.tr", "sezer.sevgin@trakyayatirim.com.tr", "uzuner.berkay1@gmail.com", "murat.aktan@trakyayatirim.com.tr", "ayca.akinci@trakyayatirim.com.tr", "sude.kinik@trakyayatirim.com.tr"]
    subject = "Yapay Zeka Newsletter"

    # Sort the matching URLs by date
    matching_urls.sort(key=lambda x: x[3])  # Assuming date is in the fourth position of each URL tuple

    # Create the email body with the sorted URLs
    body = "<html><body><div style='background-color: #0072c6; padding: 10px;'>"
    body += "<h2 style='color: #fff; margin: 0;'>Latest News:</h2>"
    body += "</div><ul style='list-style-type: none; padding: 0;'>"
    for url_info in matching_urls:
        summary, title, date, name, keywords, url = url_info
        body += f"<li style='margin-bottom: 10px;'><strong><a href=\"{url}\" style='color: #0072c6; text-decoration: none;'>{title}</a></strong><br>"
        body += f"<span style='font-size: 14px;'><strong>Date:</strong> {date}</span><br>"
        body += f"<span style='font-size: 14px;'><strong>Summary:</strong> {summary}</span><br>"
        body += f"<span style='font-size: 14px;'><strong>Keywords:</strong> {keywords.replace(selected_operator, selected_operator_formatting(selected_operator))}</span><br><br></li>"

    body += "</ul></body></html>"

    em = EmailMessage()
    em["From"] = sender_email
    em["To"] = ", ".join(reciever_emails)  # Combine recipient emails with commas

    em["Subject"] = subject
    em.set_content(body, subtype="html")
    #em.add_alternative(body, subtype="html")
    file_path = "/tmp/News.xlsx"
    # Attach the Excel file to the email
    with open(file_path, "rb") as excel_file:
        em.add_attachment(
            excel_file.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="News.xlsx"
        )

    context = ssl.create_default_context()

    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as smtp:
        smtp.login(sender_email, password)
        smtp.sendmail(sender_email, reciever_emails, em.as_string())
    
    print("email sent.")


def date_transformer(s, date_format):
    keywords = ["önce", "ago", "minute", "dakika"]

    # Check if any keyword is present in the input string
    if any(keyword in s for keyword in keywords):
        current_date = datetime.date.today()
        date = current_date.strftime("%Y/%m/%d")
        return date

    s = remove_chars_before_first_number(s)

    # ilk format 22 ağustos 2023 şeklinde
    date_list = s.split()
    if len(date_list) > 1:
        pass
    else:
        return convert_to_dd_mm_yyyy(s, date_format)
    day = date_list[0]
    year = date_list[2]
    month = 0
    month_str = date_list[1].lower()

    if month_str == "ocak":
        month = "01"
    elif month_str == "şubat":
        month = "02"
    elif month_str == "mart":
        month = "03"
    elif month_str == "nisan":
        month = "04"
    elif month_str == "mayıs":
        month = "05"
    elif month_str == "haziran":
        month = "06"
    elif month_str == "temmuz":
        month = "07"
    elif month_str == "ağustos":
        month = "08"
    elif month_str == "eylül":
        month = "09"
    elif month_str == "ekim":
        month = "10"
    elif month_str == "kasım":
        month = "11"
    elif month_str == "aralık":
        month = "12"
    else:
        return convert_to_dd_mm_yyyy(s, date_format)

    return convert_to_dd_mm_yyyy(year + "/" + month + "/" + day, 'yyyy/mm/dd')


def append_to_excel(file_path, data):
    workbook = openpyxl.load_workbook(file_path)

    # Select the first sheet (you can modify this if needed)
    sheet = workbook.active

    # Insert rows starting from the specified row
    start_row = find_next_available_row(sheet)
    
    sheet.insert_rows(start_row, len(data))

    # Write data to the inserted rows
    for row_idx, row_data in enumerate(data, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    # Save the modified workbook
    workbook.save(file_path)


def find_urls_with_keywords_or(keywords, string_url_pairs, selected_operator):
    matching_urls = []
    count = 1
    w_keywords = ""
    for i, word in enumerate(keywords):
        w_keywords += word
        if i < len(keywords) - 1:
            w_keywords += ", "
    if selected_operator == "OR":
        for string, url, name, date, title in string_url_pairs:
            for keyword in keywords:
                if " " + keyword.lower() in string.lower():
                    print(string+"***************"+"or")
                    matching_urls.append([summarizer(string),title, date, name, w_keywords.replace(",", " " + selected_operator), url])
                    break  # Stop checking other keywords for this string
    else:
        for string, url, name, date, title in string_url_pairs:

            if check_words_in_string(keywords, string):
                print(string+"***************" +"and")
                matching_urls.append([summarizer(string),title, date, name, w_keywords.replace(",", " " + selected_operator), url])

    return matching_urls


def find_matching_urls_with_keywords(keywords, selected_operator):
    keywords = str(keywords)
    download_file_from_s3("yzf1newsletterbucket", "haber_db", "/tmp/db.json")

    # Get the older links that are already processed.
    try:
        with open("/tmp/db.json", "r") as file:
            db = json.load(file)
    except FileNotFoundError:
        # If the file doesn't exist, create an empty dictionary
        db = {}
    # Get the list of links
    
    my_set = set(db.get(keywords + selected_operator, set()))
    print(len(my_set), type(my_set))

    link_list1 = general_news_links("https://webrazzi.com", my_set, "post-title", attributes="a",
                                   exclusion_list="iletisim")

    link_list2 = general_news_links("https://egirisim.com", my_set, "entry-title td-module-title",
                                    attributes="a")

    link_list3 = general_news_links("https://swipeline.co", my_set, "post-link", exclusion_list="podcast")

    link_list4 = general_news_links("https://www.girisimhaberleri.com", my_set, "post-item-title", attributes="a")

    link_list5 = general_news_links("https://www.pazarlamasyon.com", my_set, "binduz-er-trending-news-title", tag="h5",
                                    attributes="a", root_relative=True)

    # Write the set to a txt file to ensure that we don't waste computation reading
    set_to_write = list(link_list1 | link_list2 | link_list3 | link_list4 | link_list5)

    if keywords + selected_operator in db:
        db[keywords + selected_operator].extend(set_to_write)
    else:
        db[keywords + selected_operator] = set_to_write

    with open("/tmp/db.json", "w") as file:
        json.dump(db, file)
    #bunu sona al
    upload_file_to_s3("/tmp/db.json", "yzf1newsletterbucket", "haber_db")
    # Create a 2D list to store content and URLs
    content_url_list = []

    # Scrape content from each link and store it in the 2D list
    for link in link_list1:
        content, url, date, title = extract_content_from_website(link, webrazzi_rules)
        if title != 0:
            content_url_list.append([content, url, "webrazzi", date, title])
        else:
            continue
        print(date)
    print("webrazzi done")
    for link in link_list2:
        content, url, date, title = extract_content_from_website(link, egirisim_rules)
        if title != 0:
            content_url_list.append([content, url, "egirisim", date, title])
        else:
            continue
        print(date)
    print("egirişim done")
    for link in link_list3:
        content, url, date, title = extract_content_from_website(link, swipeline_rules)
        if title != 0:
            content_url_list.append([content, url, "swipeline", date, title])
        else:
            continue
        print(date)
    print("swipeline done")
    for link in link_list4:
        content, url, date, title = extract_content_from_website(link, gh_rules)
        if title != 0:
            content_url_list.append([content, url, "girişimhaberleri", date, title])
        else:
            continue
        print(date)
    print("gh done")
    for link in link_list5:
        
        content, url, date, title = extract_content_from_website(link, pazarlamasyon_rules)
        if title != 0:
            content_url_list.append([content, url, "pazarlamasyon", date, title])
        else:
            continue
        
    print("pazarlamasyon done")

    # Find URLs with keywords in the content
    matching_urls = find_urls_with_keywords_or(eval(keywords), content_url_list, selected_operator)
    download_file_from_s3("yzf1newsletterbucket", "newsletter.xlsx", "/tmp/News.xlsx")
    append_to_excel("/tmp/News.xlsx", matching_urls)
    upload_file_to_s3("/tmp/News.xlsx", "yzf1newsletterbucket", "newsletter.xlsx")
    if len(matching_urls) == 0:
        pass
    else:
        email_sender(matching_urls, selected_operator)

    return matching_urls


def lambda_handler(event, context):
    # Define your keywords and selected_operator here
    keywords = ["yapay zeka", "girişim"]
    selected_operator = "AND"

    # Your existing code for web scraping and processing goes here...
    matching_urls = find_matching_urls_with_keywords(keywords, selected_operator)

    # Return the matching URLs as a JSON response
    response = {
        "statusCode": 200,
        "body": json.dumps("success")
    }

    return response



